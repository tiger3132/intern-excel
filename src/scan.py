#Importing mysql connector
import mysql.connector
#Importing REGEX
import re

#Use regex for substitution
from re import sub

#Importing excel
import openpyxl
    
#Importing operating system
from subprocess import run

#Import date
from datetime import datetime

#Import alignment from openpyxl
from openpyxl.styles import Alignment

#Import time
import time
start_time = time.time()

import os

#Importing operating system
import subprocess

#Get current date in 'yyyy/mm/dd' format
current_date = datetime.today().strftime('%Y/%m/%d')

#Input custom format on an excel cell
CUSTOM_FORMAT = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
ACCOUNTING_FORMAT = '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-'
NUMBER_FORMAT = '#,##0'
GENERAL_FORMAT = 'General'
DATE_FORMAT = 'yyyy/mm/dd'
        
path = "/opt/PplaProject/clean.xlsx"

#Function
"""
Description: Check whether it is a holiday or not. Does not check weekends!
"""
def check_day():
    with open("day_check.txt", "w+") as f:
        
        #Put the output of the subprocess into the text file
        process = run(['ssh','heroG@pMST-01','. .profile ; ./ap/chk/redis/v_isholiday', datetime.today().strftime('%Y'), datetime.today().strftime('%Y%m%d')], stdout=f, stderr=f)
        
        #So the pointer resets to the beginning of the file
        f.seek(0)

        for x in f:
            #Pick lines only starting with #
            if x.startswith('#1'):
                print("It is a holiday")
                os._exit(os.EX_OK) 
            else:
                continue

"""
Description: Only CLEAN excel file. Cannot use max_row on file that is written on by user manually
Param: excel sheet1 variable name
"""
def new_row(sheet1):
    #Only CLEAN excel file
    #Cannot use max_row on file that is written on by user manually
    print(sheet1.max_row)
    row_write = sheet1.max_row + 1
    print(row_write)
    return row_write
    
"""
Description: Add the current date to row A sheet1
Param: excel sheet1 variable name, latest row for write variable name
"""
def add_date(sheet1, row_write):
    date = sheet1.cell(row=row_write, column=1)
    date.value = current_date
    date.number_format = DATE_FORMAT
    date.alignment = Alignment(horizontal='center')
    
    for x in range(0,7):
        check_date = sheet1.cell(row=row_write - x, column=1)
        print(str(check_date.value))
    
"""
Description: Create excel file
"""
def excel_create():
    #Reading and writing to excel file at path

    #Enable writing on an excel sheet
    wb_obj = openpyxl.load_workbook(path)

    #Access each sheets in an excel file
    sheet1 = wb_obj.get_sheet_by_name("History Data")

    sheet2 = wb_obj.get_sheet_by_name("Server")
    
    return sheet1, sheet2, wb_obj
    
"""
Description: Writing into the excel sheet
params: sheet, row number, column number, value of cell, format of number
"""
def excel_write(sheet, row_num, col_num, value, num_format = GENERAL_FORMAT):
    a = sheet.cell(row=row_num, column=col_num)
    a.value = value
    a.number_format = num_format

    
"""
Description: Total Connection User Count in Sheet1 and Sheet2 (Except Auto). Sheet1 cell: B - F (No D). Sheet2 cell: 9D - 14D (No 12D)
Param: excel sheet1 variable name, excel sheet2 variable name, latest row for write variable name
"""
def total_connect_user_count(sheet1, sheet2, row_write):
    #HTS:
    #Sheet1
    mycursor.execute("SELECT DISTINCT(USER_ID) FROM HEROG.MS_CUST_CNCT_RECD WHERE CNCT_TIME > CONCAT(CURDATE(), ' 00:00:00') AND CNCT_TIME <= CONCAT(CURDATE(), ' 18:00:00') AND CNCT_MDIA_TP = '01'")
    
    excel_write(sheet1, row_write, 2, mycursor.rowcount, CUSTOM_FORMAT)
    
    #Sheet2
    
    excel_write(sheet2, 9, 4, mycursor.rowcount, ACCOUNTING_FORMAT)

    #MTS:
    #Sheet1
    mycursor.execute("SELECT DISTINCT(USER_ID) FROM HEROG.MS_CUST_CNCT_RECD WHERE CNCT_TIME > CONCAT(CURDATE(), ' 00:00:00') AND CNCT_TIME <= CONCAT(CURDATE(), ' 18:00:00') AND CNCT_MDIA_TP = '02'")

    excel_write(sheet1, row_write, 3, mycursor.rowcount, CUSTOM_FORMAT)
    
    #Sheet2
    excel_write(sheet2, 10, 4, mycursor.rowcount, CUSTOM_FORMAT)    

    #TRADER:
    #Sheet1
    mycursor.execute("SELECT DISTINCT(USER_ID) FROM HEROG.MS_CUST_CNCT_RECD WHERE CNCT_TIME > CONCAT(CURDATE(), ' 00:00:00') AND CNCT_TIME <= CONCAT(CURDATE(), ' 18:00:00') AND CNCT_MDIA_TP = '03'")

    excel_write(sheet1, row_write, 5, mycursor.rowcount, CUSTOM_FORMAT)
    
    #Sheet2
    excel_write(sheet2, 14, 4, mycursor.rowcount, ACCOUNTING_FORMAT)
    
    #TOTAL:
    #Sheet1
    mycursor.execute("SELECT DISTINCT(USER_ID) FROM HEROG.MS_CUST_CNCT_RECD WHERE CNCT_TIME > CONCAT(CURDATE(), ' 00:00:00') AND CNCT_TIME <= CONCAT(CURDATE(), ' 18:00:00')")

    excel_write(sheet1, row_write, 6, mycursor.rowcount, CUSTOM_FORMAT)
    
    #Sheet2
    excel_write(sheet2, 15, 4, mycursor.rowcount, CUSTOM_FORMAT) 
    
"""
Description: Concurrent User in Sheet1 and Sheet2 (Except Auto). Sheet1 cell: G - K (No I). Sheet2 cell: 9E - 14E (No 12E)
Param: excel sheet1 variable name, excel sheet2 variable name, latest row for write variable name
"""    
def concurrent_user(sheet1, sheet2, row_write):
    #Concurrent users text file
    f_concur = open("/var/www/html/mon/output_ex1_1", "r")

    line = f_concur.readline()
    #Split by non-characters
    values = re.split(r'[^\w]+',line)

    #CONCURRENT USERS SUM:
    #Sheet1
    excel_write(sheet1, row_write, 11, int(values[1]), CUSTOM_FORMAT)
    
    #Sheet2
    excel_write(sheet2, 15, 5, int(values[1]), ACCOUNTING_FORMAT)
    
    #HTS:
    #Sheet1
    excel_write(sheet1, row_write, 7, int(values[3]), CUSTOM_FORMAT)
    
    #Sheet2
    excel_write(sheet2, 9, 5, int(values[3]), ACCOUNTING_FORMAT)
    
    #MTS:
    #Sheet1
    excel_write(sheet1, row_write, 8, int(values[5]), CUSTOM_FORMAT)
    
    #Sheet2
    excel_write(sheet2, 10, 5, int(values[5]), ACCOUNTING_FORMAT)
    
    #TRADER:
    #Sheet1
    excel_write(sheet1, row_write, 10, int(values[7]), CUSTOM_FORMAT)

    #Sheet2
    excel_write(sheet2, 14, 5, int(values[7]), ACCOUNTING_FORMAT)

"""
Description: User with Order Submission in Sheet1 and Sheet2. Sheet1 cell: L - N. Sheet2 cell: 9F -14F
Param: excel sheet1 variable name, excel sheet2 variable name, latest row for write variable name
"""
def user_order_submission(sheet1, sheet2, row_write):

    #Order submission text file
    f_ordsub = open("/var/www/html/mon/output_ex1_5", "r")

    #List for all the values that are needed for excel
    list_of_val = []

    #Looping through the file
    for line in f_ordsub:
        #Skipping the files with empty strings
        if line == '\n':
            continue
        else:
            #Used N for split to two values in list
            x = line.split('N')
            #Use first value
            #Get the values inside the ()  
            searchObj = re.search('(?<=\()(.*?)(?=\))',x[1], re.M|re.I)
            list_of_val.append(int(searchObj.group()))
    
    #HTS
    #Sheet1
    excel_write(sheet1, row_write, 12, list_of_val[0])
    
    #Sheet2
    excel_write(sheet2, 9, 6, list_of_val[0], ACCOUNTING_FORMAT)

    #MTS
    #Sheet1
    excel_write(sheet1, row_write, 13, list_of_val[1], CUSTOM_FORMAT)

    #Sheet2
    excel_write(sheet2, 10, 6, list_of_val[1], CUSTOM_FORMAT)

    #TRADER
    #Sheet1
    excel_write(sheet1, row_write, 14, list_of_val[2], CUSTOM_FORMAT)
    
    #Sheet2
    excel_write(sheet2, 14, 6, list_of_val[2], ACCOUNTING_FORMAT)
    
"""
Description: Auto for Total Connection User and Concurrent User in Sheet1 and Sheet2. Sheet1 cell: D and I. Sheet2 cell: 12D and 12E
Param: excel sheet1 variable name, excel sheet2 variable name, latest row for write variable name
"""
def auto_user(sheet1, sheet2, row_write):
    #Sheet1
    connect_user = 0
    concurr_user = 0
    #Writing into the text file then reading it
    with open("auto.text", "w+") as f:
        
        #Put the output of the subprocess into the text file
        process = run(['ssh','heroG@pMST-01','. .profile ; ./ap/ssao/v_sao_mon_cond'], stdout=f, stderr=f)
        #process = pipesub.run(["ssh","heroG@pMST-01",". .profile ; ./ap/ssao/v_sao_mon_cond"], stdout=pipesub.PIPE, stderr=pipesub.PIPE)
        #ssh heroG@pMST-01 ". .profile ; ./ap/ssao/v_sao_mon_cond"
        
        #Reset the pointer of reading the text file
        f.seek(0)
        
        for x in f:
            if x.startswith('#'):
                b = x.split('|')
                b[6] = sub('[\n]', '', b[6])
                #Sum the values
                connect_user = connect_user + int(b[1]) + int(b[2]) + int(b[3])
                concurr_user = concurr_user + int(b[4]) + int(b[5]) + int(b[6])
            #Format with coma
            #Using excel number formatter does not work
            value_connect = "{:,d}\n{}".format(connect_user, "(cond. set)")
            value_concurr = "{:,d}\n{}".format(concurr_user, "(cond. started)")
            
            #Sheet1
            #Connection Users
            excel_write(sheet1, row_write, 4, connect_user, CUSTOM_FORMAT)
            #Concurrent Users
            excel_write(sheet1, row_write, 9, concurr_user, CUSTOM_FORMAT)
            
            #Sheet2
            #Connection Users
            excel_write(sheet2, 12, 4, value_connect)
            #Concurrent Users
            excel_write(sheet2, 12, 5, value_concurr)

"""
Description: User Count for HTS mode in Sheet1. Sheet1 cell: O - S
Param: excel sheet1 variable name, latest row for write variable name
"""
def user_count_HTS_mode(sheet1, row_write):

    #File for HTS mode
    f_htsmode = open("/var/www/html/mon/output_ex1_4", "r")

    #Pick specific lines in the file
    count = 1

    #Values needed for excel
    list_of_val = []

    #Looping through the file
    for line in f_htsmode:
        if count%2 == 0:
            list_of_val.append(int(line))
        count = count + 1
    
    #Sheet1
    #Beginner
    excel_write(sheet1, row_write, 16, list_of_val[0])
    
    #Standard
    excel_write(sheet1, row_write, 17, list_of_val[1], CUSTOM_FORMAT)
    
    #Compact
    excel_write(sheet1, row_write, 18, list_of_val[2])
    
    #Mini
    excel_write(sheet1, row_write, 19, list_of_val[3])
    
    #Classic
    excel_write(sheet1, row_write, 15, list_of_val[5])

"""
Description: Download Count for HTS in Sheet1 and Sheet2. Sheet1 cell: AE. Sheet2 cell: 9G
Param: excel sheet1 variable name, excel sheet2 variable name, latest row for write variable name
"""
def download_count(sheet1, sheet2, row_write):
    
    #File for download_count
    f_down = open("/var/www/html/mon/output_ex1_3", "r")
    
    line = f_down.readline()
    
    #HTS Downloads
    #Sheet1
    excel_write(sheet1, row_write, 31, int(line))
    
    #Sheet2
    excel_write(sheet2, 9, 7, int(line), ACCOUNTING_FORMAT)
    
"""
Description: No. of Smart access login in Sheet1. Sheet1 cell: BE
Param: excel sheet1 variable name, latest row for write variable name
"""
def smart_login(sheet1, row_write):
    
    #File for download_count
    f_smart = open("/var/www/html/mon/output_ex1_2", "r")
    
    line = f_smart.readline()
    #Total login
    excel_write(sheet1, row_write, 57, int(line), CUSTOM_FORMAT)
    
"""
Description: User connection count in Sheet2. Sheet2 cell: 19C - 19G, 19J, 19K
Param: excel sheet2 variable name, latest row for write variable name
"""
def user_connect_count(sheet2, row_write):
    
    #MTS Mode
    mycursor.execute("SELECT android.COUNT AS android, ios.COUNT AS iphone FROM(SELECT DATE(a.CNCT_TIME) AS DATE, COUNT(DISTINCT a.USER_ID) AS COUNT FROM HEROG.MS_CUST_CNCT_RECD a WHERE a.MBL_PLTF_TP = '01' GROUP BY DATE(a.CNCT_TIME)) AS android LEFT JOIN( SELECT DATE(a.CNCT_TIME) AS DATE, COUNT(DISTINCT a.USER_ID) AS COUNT FROM HEROG.MS_CUST_CNCT_RECD a WHERE a.MBL_PLTF_TP = '02' OR a.MBL_PLTF_TP = '03' GROUP BY DATE(a.CNCT_TIME)) AS ios ON android.DATE = ios.DATE ORDER BY android.DATE DESC LIMIT 1")

    mobile_result = mycursor.fetchall()
    
    #For [][] the first index is the row and the second index is tuple's index
    #If a value is already in the cell, then this statement will replace it
    
    #Android
    excel_write(sheet2, 19, 10, mobile_result[0][0], CUSTOM_FORMAT)
    
    #iPhone
    excel_write(sheet2, 19, 11, mobile_result[0][1], CUSTOM_FORMAT)
    
    
    #HTS Mode
    mycursor.execute("SELECT Standard.DATE, Beginner.COUNT AS Beginner, Standard.COUNT AS Standard, Compact.COUNT AS Compact, Mini.COUNT AS Mini, Classic.COUNT AS Classic FROM( SELECT DATE, COUNT FROM FSS.SCREEN_MODE WHERE MODE = 1) AS Standard LEFT JOIN( SELECT DATE, COUNT FROM FSS.SCREEN_MODE WHERE MODE = 0) AS Beginner ON Beginner.DATE = Standard.DATE LEFT JOIN( SELECT DATE, COUNT FROM FSS.SCREEN_MODE WHERE MODE = 2) AS Compact ON Compact.DATE = Standard.DATE LEFT JOIN( SELECT DATE, COUNT FROM FSS.SCREEN_MODE WHERE MODE = 3) AS Mini ON Mini.DATE = Standard.DATE LEFT JOIN( SELECT DATE, COUNT FROM FSS.SCREEN_MODE WHERE MODE = 5) AS Classic ON Classic.DATE = Standard.DATE ORDER BY 1 DESC LIMIT 1")
    
    hts_result = mycursor.fetchall()
    #Classic
    excel_write(sheet2, 19, 3, hts_result[0][5], CUSTOM_FORMAT)
    
    #Beginner
    excel_write(sheet2, 19, 4, hts_result[0][1], CUSTOM_FORMAT)
    
    #Standard
    excel_write(sheet2, 19, 5, hts_result[0][2], CUSTOM_FORMAT)
    
    #Compact
    excel_write(sheet2, 19, 6, hts_result[0][3], CUSTOM_FORMAT)
    
    #Mini
    excel_write(sheet2, 19, 7, hts_result[0][4], CUSTOM_FORMAT)

"""
Description: Matched Amount of different user connection in Sheet1 and Sheet2. Sheet1 cell: U - AC. Sheet2 cell: 19C - 19G, 19J, 19K. (Matched Amount is the same as Trading Value)
Data other than overnight comes after 18:30
Param: excel sheet1 variable name, excel sheet2 variable name, latest row for write variable name
"""
def matched_amount(sheet1, sheet2, row_write):  
#/opt/PplaProject/test.txt
    #File for matched amount
    with open("/var/www/html/mon/output_ex1_6", "r") as f:
        
        line = 1
        for x in f:
            b = x.split(':')
            
            if line == 2:
                #Get rid of '\n' at the end of the line
                #and ',' inside the number
                b[2] = sub('[\n,]', '', b[2])
                #Trading Value MTS - Overnight
                #Sheet1
                excel_write(sheet1, row_write, 21, int(b[2]), NUMBER_FORMAT)
            elif line == 3:
                b[2] = sub('[\n,]', '', b[2])
                #Trading Value MTS - Android
                #Sheet1
                excel_write(sheet1, row_write, 22, int(b[2]), NUMBER_FORMAT)
                #Sheet2
                excel_write(sheet2, 21, 10, int(b[2]), NUMBER_FORMAT)
            elif line == 4:
                b[2] = sub('[\n,]', '', b[2])
                #Trading Value MTS - iPhone
                #Sheet1
                excel_write(sheet1, row_write, 23, int(b[2]), NUMBER_FORMAT)
                
                #Sheet2
                excel_write(sheet2, 21, 11, int(b[2]), NUMBER_FORMAT)
            elif line == 6:
                b[2] = sub('[\n,]', '', b[2])
                #Trading Value HTS - Overnight
                #Sheet1
                excel_write(sheet1, row_write, 24, int(b[2]), NUMBER_FORMAT)
               
            elif line == 7:
                b[2] = sub('[\n,]', '', b[2])
                #Trading Value HTS - Beginner
                #Sheet1
                excel_write(sheet1, row_write, 26, int(b[2]), NUMBER_FORMAT)
                
                #Sheet2
                excel_write(sheet2, 21, 4, int(b[2]), ACCOUNTING_FORMAT)
            elif line == 8:
                b[2] = sub('[\n,]', '', b[2])
                #Trading Value HTS - Standard
                #Sheet1
                excel_write(sheet1, row_write, 27, int(b[2]), NUMBER_FORMAT)
                
                #Sheet2
                excel_write(sheet2, 21, 5, int(b[2]), ACCOUNTING_FORMAT)
            elif line == 9:
                b[2] = sub('[\n,]', '', b[2])
                #Trading Value HTS - Compact
                #Sheet1
                excel_write(sheet1, row_write, 28, int(b[2]), NUMBER_FORMAT)
                
                #Sheet2
                excel_write(sheet2, 21, 6, int(b[2]), ACCOUNTING_FORMAT)
            elif line == 10:
                b[2] = sub('[\n,]', '', b[2])
                #Trading Value HTS - Mini
                #Sheet1
                excel_write(sheet1, row_write, 29, int(b[2]), NUMBER_FORMAT)
                
                #Sheet2
                excel_write(sheet2, 21, 7, int(b[2]), ACCOUNTING_FORMAT)
            elif line == 11:
                b[2] = sub('[\n,]', '', b[2])
                #Trading Value HTS - Classic
                #Sheet1
                excel_write(sheet1, row_write, 25, int(b[2]), NUMBER_FORMAT)
            
                #Sheet2
                excel_write(sheet2, 21, 3, int(b[2]), ACCOUNTING_FORMAT)
            
            line = line + 1     

"""
Description: Calculating sum with different cells in Sheet1. Sheet1 cell: T, AD, AF, AH, AJ
Param: excel sheet1 variable name, latest row for write variable name
"""
def calculate_sum(sheet1, row_write):
    #User count for HTS
    excel_write(sheet1, row_write, 20, "=SUM(O{}:S{})".format(row_write, row_write), CUSTOM_FORMAT)
    #Matched Amount for HTS
    excel_write(sheet1, row_write, 30, "=SUM(X{}:AC{})".format(row_write, row_write), NUMBER_FORMAT)
    #Download count HTS
    excel_write(sheet1, row_write, 32, "=AF{}+AE{}".format(row_write-1, row_write), CUSTOM_FORMAT)
    #Download count iphone
    excel_write(sheet1, row_write, 34, "=AH{}+AG{}".format(row_write-1, row_write), CUSTOM_FORMAT)
    #Download count android
    excel_write(sheet1, row_write, 36, "=AJ{}+AI{}".format(row_write-1, row_write), CUSTOM_FORMAT)

"""
Description: Have to save excel file so changes are shown
Param: Excel workbook variable
"""
def save_file(wb_obj):
    wb_obj.save(path)

#Connecting to the database of Finansia Syrus
mydb = mysql.connector.connect(
  host="172.21.9.11",
  user="fss",
  password="fss@dmin",
  database="HEROG"
)

#Used to execute statements to communicate with the MySQL database
mycursor = mydb.cursor(buffered=True)

"""
for cell in sheet1["A"]:
    
    
    if cell.row == 2 or cell.row == 3 or cell.row == 4:
        continue
    
    elif cell.value is None:
        row_write2 = cell.row
        print("ROW = {}".format(row_write2))
        break
"""

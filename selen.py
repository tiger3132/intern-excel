#Used for interacting with excel file
import openpyxl
#Adding font
from openpyxl.styles import Font

#Importing webdriver for Firefox
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
#Specific exception
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import NoSuchElementException

from selenium.webdriver.common.keys import Keys

#Importing time
from time import sleep
import time
start_time = time.time()

import os

#Importing operating system
import subprocess

#Import date
from datetime import datetime

#For sending errors and miscelaneous
import mail

#Regex
from re import split

#Input custom format on an excel cell
GENERAL_FORMAT = 'General'

#Functions
"""
Description: Screenshot of firefox webdriver
Param: number (added to screenshot file name)
"""
def screenshot(screen_num):
    print(driver.get_screenshot_as_file('./foo'+str(screen_num)+'.png'))

#Sleep for 2 seconds
#Does not require sleep for 2 seconds because the function takes long enough to make the element not stale
def wait_for(condition_function):
    if condition_function:
        return True
    else:
        sleep(2)
        print("Timeout for 2s")
        return False

#Check if link is stale or not
# def click_link_at_new_page(link_text):
    # start_time = time.time()
    # wait = False
    # while wait == False:
        # def link_check():
            # try:
                # # poll the link with an arbitrary call
                # link = driver.find_element_by_partial_link_text(link_text)
                # link.click()
                # return True
            # except StaleElementReferenceException:
                # return False
        # link = link_check()
        # wait = wait_for(link)
        # print(f'{time.time() - start_time:.2f}s')
        
#An error comes up and early quit
def error():
    print('Check screenshot for more detail')
    driver.close()
    driver.quit()
    subprocess.run(["pkill","-f","firefox"])
    os._exit(os.EX_OK) 

"""
Description: When click() gives exception
Param: the specific exception function eg. link_has_gone_stale, the element eg. link name, xpath, css
"""
def click_element(check_exception, element):
    wait = False
    print(1)
    #Loop until there is no more exception
    while wait == False:
        print(2)
        has_exception = check_exception(element)
        print(5)
        wait = wait_for(has_exception)
        print(6)
         
#Stale exception when click on partial link
def link_has_gone_stale(element):
    try:
        print(3)
        link = driver.find_element_by_partial_link_text(element)
        link.click()
        print('Not Stale')
        return True
    except StaleElementReferenceException:
        print(4)
        print('Stale')
        return False  
#Click intercepted exception when click on xpath
def xpath_block(element):
    try:
        link = driver.find_element_by_xpath(element)
        link.click()
        return True
    except ElementClickInterceptedException:
        return False
        
"""
Description: Writing into the excel sheet
Param: sheet, row number, column number, value of cell, format of number
"""
def excel_write(sheet, row_num, col_num, value, num_format = GENERAL_FORMAT):
    a = sheet.cell(row=row_num, column=col_num)
    a.value = value
    a.number_format = num_format

"""
Description: Selenium process of Finansia's google app to obtain Android download count. Sheet1: AI. Sheet2: J19
Param: excel sheet1 variable name, excel sheet2 variable name, latest row for write variable name
"""
def google(sheet1, sheet2, row_write):
    count = 0
    while count < 7:
        #link for google play console of Finansia
        driver.get('https://accounts.google.com/signin/v2/identifier?service=androiddeveloper&passive=1209600&continue=https%3A%2F%2Fplay.google.com%2Fapps%2Fpublish%2F%3Faccount%3D7543871693245328704%23AppListPlace&followup=https%3A%2F%2Fplay.google.com%2Fapps%2Fpublish%2F%3Faccount%3D7543871693245328704&flowName=GlifWebSignIn&flowEntry=ServiceLogin&hl=en')
        #Logging in

        try:
            email = driver.find_element_by_id("identifierId")
            email.send_keys(app_mail)
        except:
            send_error()
            screenshot(1)
            mail.send_error("ERROR: Cannot find the input box for email/Source page havent load fully")
            
        captcha = driver.find_element_by_id("ca")


        #When a captcha is found when logging in
        if captcha.is_displayed():
            screenshot(2)
            x = input("Write the CAPTCHA: ")
            captcha.send_keys(x)

        try:
            enter_button = driver.find_element_by_xpath("//div[@id='identifierNext']/div/button")
            enter_button.click()
        except:
            screenshot(3)
            mail.send_error('ERROR: Cannot find enter button/Captcha is wrong')
            
        #Pausing for the website to load in seconds
        sleep(5)

        #Logging in password
        try:
            enter_password = driver.find_element_by_xpath("//div[@id='password']/div/div/div/input")
            enter_password.send_keys(app_pass)
        except:
            screenshot(4)
            mail.error('Cannot find input box for password/Page is not fully loaded')
            
            
        password_button = driver.find_element_by_xpath("//div[@id='passwordNext']/div/button")
        password_button.click()
        #Only do the commented code below when logging in the for the first time
        """
        sleep(10)

        recovery_email = driver.find_element(By.XPATH,("//div[@class='vxx8jf']"))

        recovery_email.click()

        sleep(5)

        recover_authen = driver.find_element(By.XPATH,("//div[@class='Xb9hP']/input"))
        recover_authen.send_keys("fssplaystore@gmail.com")

        button_authen = driver.find_element(By.XPATH,("//div[@class='qhFLie']/div/div/button"))
        button_authen.click()

        sleep(10)                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                           
        print(driver.get_screenshot_as_file('./foo6.png'))
        """

        #Choose Finansia Hero app

        all_apps_load = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//div[@data-title-type='SECTION']/div[1]/button")))

        #Sometimes give a StaleElement exception 
        try:
            click_element(link_has_gone_stale, "kr.co.daou.fss")
            break
        #A different google page shows up because of suspicious activity
        #Create new tab until this page does not show up
        except NoSuchElementException:
            mail.send_error("Unable to locate element kr.co.daou.fss")
            screenshot(5)
            f = open("demofile10.txt", "w")
            f.write(driver.page_source)
            f.close()
            #Creates a new tab and redo
            driver.find_element_by_tag_name('body').send_keys(Keys.CONTROL + 't')
            count = count + 1

    if count == 7:
        mail.send_error('ERROR: Unable to locate element kr.co.daou.fss after 7 retries')
        error()
        
    sleep(2)
    #Getting statistics
    stats_link = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//a[@title='Statistics']"))).click()

    print('success')
    sleep(5)

    #Changing the time period to life time
    try:
        time_period = driver.find_element_by_xpath("//section/div[2]/div/div/div/div/div[2]/div/div/button")
        time_period.click()
    except:
        screenshot(6)
        mail.send_error('Element is being blocked by another element/Element cannot be found')
        error()

    last_thirty = driver.find_element_by_xpath("//section/div[2]/div/div/div/div/div[2]/div/div/div/div/div/div/div[2]/button[2]")
    last_thirty.click()


    done_button = driver.find_element_by_xpath("//section/div[2]/div/div/div/div/div[2]/div/div/div/div/div[2]/button[2]")
    done_button.click()

    sleep(5)

    #Switching to stats about new users acquisition
    #First method 
    """
    new_users_button = driver.find_element_by_xpath("//section/div[2]/div/div/div[3]/div/button")

    new_users_button.click()

    users_hover = driver.find_element_by_xpath("//section/div[2]/div/div/div[3]/div/div/div/div[1]")

    users_hover.click()

    users_acquisition_hover = driver.find_element_by_xpath("//section/div[2]/div/div/div[3]/div/div/div[2]/div/div[2]")

    users_acquisition_hover.click()

    users_acquisition_hover = driver.find_element_by_xpath("//section/div[2]/div/div/div[3]/div/div/div[2]/div/div[2]/div[3]/div/div[2]")

    users_acquisition_hover.click()
    """
    #Second method
    #Saved report
    quick_reports = driver.find_element_by_xpath("//div[@role='article']/div/div/div/div/button")
    quick_reports.click()

    pla_report = driver.find_element_by_xpath("//div[@role='article']/div/div/div/div/div/div/div/div/div[1]")
    pla_report.click()

    #Show all rows for stats
    try:
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//section/div[3]/div[4]/div/div[3]/div[2]/div[1]"))).click()
    except:
        screenshot(4123123)
    all_rows = driver.find_element_by_xpath("//body/div[9]/div/div/div/div/div[3]")
    all_rows.click()

    sleep(5)
    
    #Variable for latest date
    latest_data_row = 0
    n = 0
    found_date = False
    date_row = driver.find_element_by_xpath("//section/div[3]/div[4]/div/div[3]/div/table/tbody/tr[last()]/td[1]/div/div/h3")
    print(date_row.text)
    #Loop through excel found the latest date with android downloads
    while found_date == False:
        print(n)
        check_date = sheet1.cell(row=row_write - n, column=1)
        print(str(check_date.value))
        #To get the individ date values in excel
        b = split('[-\/\s]', str(check_date.value))
        print(b)
        convert_date = datetime(int(b[0]), int(b[1]), int(b[2])).strftime("%a, %b %-d, %Y")
        print(convert_date)
        #Check if date of excel is same as date of android download
        if(convert_date == date_row.text):
            latest_data_row = row_write - n
            print(latest_data_row)
            found_date = True
            continue
        excel_write(sheet1, row_write - n, 35, 0, CUSTOM_FONT)
        n = n + 1
    
    last_day = 7
    #Add up downloads for holidays
    sum_data = 0 
    #y loops through the excel file
    y = 0
    #x loops through the last 7 days of website
    for x in range(0,last_day):
        print(y)
        date_stuff = sheet1.cell(row=latest_data_row - y, column=1)
        print(date_stuff.value)
        c = split('[-\/\s]', str(date_stuff.value))
        print(c)
        convert_date = datetime(int(c[0]), int(c[1]), int(c[2])).strftime("%a, %b %-d, %Y")
        single_row = driver.find_element_by_xpath("//section/div[3]/div[4]/div/div[3]/div/table/tbody/tr[last()-"+str(x)+"]/td[2]/div/div/h3")
        single_date = driver.find_element_by_xpath("//section/div[3]/div[4]/div/div[3]/div/table/tbody/tr[last()-"+str(x)+"]/td[1]/div/div/h3")
        #Check if the day is a holiday or not. If it is, then don't write to excel but accumulate until reach a non-holiday. Then add it to that day.
        if single_date.text != convert_date:
            #num variable to hold the date where to put the holiday variables
            hold = y
            print(hold)
            sum_data = sum_data + int(single_row.text)
            write_holiday_values = driver.find_element_by_xpath("//section/div[3]/div[4]/div/div[3]/div/table/tbody/tr[last()-"+str(hold)+"]/td[2]/div/div/h3")
            if x >= 7:
                last_day = last_day + 1
            sum_data = sum_data + int(single_row.text)
            print(f'sum data = {sum_data}')
            continue
        elif single_row == '-':
            excel_write(sheet1, latest_data_row - y, 35, 0, CUSTOM_FONT)
            y = y + 1
            continue
            
        excel_write(sheet1, latest_data_row - y, 35, int(single_row.text) + sum_data, CUSTOM_FONT)
        if sum_data > 0:
            excel_write(sheet1, latest_data_row - hold, 35, int(write_holiday_values.text) + sum_data)
        sum_data = 0
        y = y + 1
        
        #Most recent value for sheet2
        last_row = driver.find_element_by_xpath("//section/div[3]/div[4]/div/div[3]/div/table/tbody/tr[last()]/td[2]/div/div/h3")

        excel_write(sheet2, 11, 7, int(last_row.text), ACCOUNTING_FONT)
    """    
    #Most recent value for sheet2
    last_row = driver.find_element_by_xpath("//section/div[3]/div[4]/div/div[3]/div/table/tbody/tr[last()]/td[2]/div/div/h3")

    excel_write(sheet2, 11, 7, int(last_row.text), ACCOUNTING_FONT)
    """
    """
    #Write the values into excel 
    #Android downloads
    for x in range(0,7):
        single_row = driver.find_element_by_xpath("//section/div[3]/div[4]/div/div[3]/div/table/tbody/tr[last()-"+str(x)+"]/td[2]/div/div/h3")
        if single_row == '-':
            excel_write(sheet1, row_write - x, 35, 0, CUSTOM_FONT)
            continue
        excel_write(sheet1, row_write - x, 35, int(single_row.text), CUSTOM_FONT)

    """
"""
Description: Selenium process of Finansia's apple app to obtain Iphone download count. Sheet1: AG. Sheet2: K19
Param: excel sheet1 variable name, excel sheet2 variable name, latest row for write variable name
"""
def apple(sheet1, sheet2, row_write):
    #Link for Finansia Syrus appstore
    driver.get('https://appstoreconnect.apple.com/login?module=AppAnalytics&hostname=analytics.itunes.apple.com&targetUrl=%2F&authResult=FAILED#%2Fapps%3Ft=all&interval=r&datesel=20171130:20200721')

    #Had to switch frames because of #document under iframe
    try:
        WebDriverWait(driver, 30).until(EC.frame_to_be_available_and_switch_to_it((By.ID,"aid-auth-widget-iFrame")))
    except:
        send_error("Cannot find iframe")
        screenshot(7)
    #Logging in to apple
    apple_id = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.ID,"account_name_text_field"))).send_keys(g_mail)

    sleep(3)

    button = driver.find_element_by_id("sign-in")
    button.click()

    apple_password = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID,"password_text_field"))).send_keys(g_pass)

    sleep(2)

    button = driver.find_element_by_id("sign-in")
    button.click()

    #If appleid has otp
    if "Enter the code to continue" in driver.page_source:
        
        a,b,c,d,e,f = input("Enter numbers: ").split()

        print(a)

        first_input = driver.find_element_by_xpath("//div[@localiseddigit='Digit']/div/div[1]/input")
        first_input.send_keys(a)

        second_input = driver.find_element_by_xpath("//div[@localiseddigit='Digit']/div/div[2]/input")
        second_input.send_keys(b)

        third_input = driver.find_element_by_xpath("//div[@localiseddigit='Digit']/div/div[3]/input")
        third_input.send_keys(c)

        fourth_input = driver.find_element_by_xpath("//div[@localiseddigit='Digit']/div/div[4]/input")
        fourth_input.send_keys(d)

        fifth_input = driver.find_element_by_xpath("//div[@localiseddigit='Digit']/div/div[5]/input")
        fifth_input.send_keys(e)

        sixth_input = driver.find_element_by_xpath("//div[@localiseddigit='Digit']/div/div[6]/input")
        sixth_input.send_keys(f)

        sleep(5)

    if "protect your information" in driver.page_source:
        privacy_button = driver.find_element_by_xpath("//div[@class='primary-button-group']/button")
        privacy_button.click()
        sleep(5)

        trust_button = driver.find_element_by_xpath("//button[starts-with(@id, 'trust-browser')]")
        trust_button.click()


    #Switch to default frame or else will not show any html/css
    driver.switch_to.default_content()

    #access finansia hero app
    app_analytics_page = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='question']")))


    finan_hero = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH,"//table[@class='dualtable']/tbody/tr[1]/td[1]/a/div"))).click()  

    #config the stats

    app_units = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,"//div[@label='App Units']/div/div[3]"))).click()

    #Wait for page to load by wait until element located
    sales_page_load = WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.XPATH, "//div[@data-key='units']/div/div/div/div")))


    per_period = driver.find_element_by_xpath("//div[@poptitle='Months']/div")
    per_period.click()

    per_day = driver.find_element_by_xpath("//div[@poptitle='Months']/div[3]/div/div[2]/div/ul/li[1]")
    per_day.click()


    time_range = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='chartdatepicker']/div/div[2]")))

    #Can give error when class hitarea obscures it
    click_element(xpath_block, "//div[@class='chartdatepicker']/div/div[2]")

    last_30 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//div[@style='height: 262px;']/div/div/div[2]")))
    last_30.click()

    sleep(5)

    #iPhone downloads
    #Sheet1
    
    #Variable for latest date
    latest_data_row = 0
    n = 0
    found_date = False
    date_row = driver.find_element_by_xpath("//table[@width='100%']/tbody/tr[1]/td[1]")
    print(date_row.text)
    #Loop through excel found the latest date with iphone downloads
    while found_date == False:
        print(n)
        check_date = sheet1.cell(row=row_write - n, column=1)
        print(str(check_date.value))
        #To get the individ date values in excel
        b = split('[-\/\s]', str(check_date.value))
        print(b)
        convert_date = datetime(int(b[0]), int(b[1]), int(b[2])).strftime("%b %-d")
        print(convert_date)
        #Check if date of excel is same as date of iphone download
        if(convert_date == date_row.text):
            latest_data_row = row_write - n
            print(latest_data_row)
            found_date = True
            continue
        excel_write(sheet1, row_write - n, 33, 0)
        n = n + 1
        
    last_day = 7
    #Add up downloads for holidays
    sum_data = 0 
    #y loops through the excel file
    y = 0
    #x loops through the last 7 days of website
    for x in range(0,last_day):
        print(y)
        date_stuff = sheet1.cell(row=latest_data_row - y, column=1)
        print(date_stuff.value)
        c = split('[-\/\s]', str(date_stuff.value))
        #['2020', '09', '10', '00:00:00']
        print(c)
        convert_date = datetime(int(c[0]), int(c[1]), int(c[2])).strftime("%b %-m")
        single_row = driver.find_element_by_xpath("//table[@width='100%']/tbody/tr[" + str(x+1) + "]/td[2]/div")
        single_date = driver.find_element_by_xpath("//table[@width='100%']/tbody/tr[" + str(x+1) + "]/td[1]")
        #Check if the day is a holiday or not. If it is, then don't write to excel but accumulate until reach a non-holiday. Then add it to that day.
        """
        if single_date.text != convert_date:
            
            sum_data = sum_data + int(single_row.text)
            print(f'sum data = {sum_data}')
            continue
        """
        if single_date.text != convert_date:
            #num variable to hold the date where to put the holiday variables
            hold = y
            print(hold)
            write_holiday_values = driver.find_element_by_xpath("//table[@width='100%']/tbody/tr[" + str(hold) + "]/td[2]/div")
            if x >= 7:
                last_day = last_day + 1
            sum_data = sum_data + int(single_row.text)
            print(f'sum data = {sum_data}')
            continue
        elif single_row == '-':
            excel_write(sheet1, latest_data_row - y, 33, 0)
            y = y + 1
            continue
            
        excel_write(sheet1, latest_data_row - y, 33, int(single_row.text))
        if sum_data > 0:
            excel_write(sheet1, latest_data_row - hold, 33, int(write_holiday_values.text) + sum_data)
        sum_data = 0
        y = y + 1
    """
    for x in range(0,7):
        single_row = driver.find_element_by_xpath("//table[@width='100%']/tbody/tr[" + str(x+1) + "]/td[2]/div")
        excel_write(sheet1, row_write - x, 33, int(single_row.text))
    """
    

    #Sheet2
    last_row = driver.find_element_by_xpath("//table[@width='100%']/tbody/tr[1]/td[2]/div")
    excel_write(sheet2, 10, 7, int(last_row.text), CUSTOM_FONT)
    
    #Make sure the driver is safely closed
    driver.close()
    driver.quit()
    subprocess.run(["pkill","-f","firefox"])

#Custom font for excel
CUSTOM_FONT = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
ACCOUNTING_FONT = '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-'

#Reading and writing to excel file
path = "/opt/PplaProject/clean.xlsx"

#appleid
app_mail = "fssfinansiahero@gmail.com"
app_pass = ""

#googleplay
g_mail = "ebizfss2@hotmail.com"
g_pass = ""

#Access webdriver option
options = Options()
#Firefox will not have any user interaction on the screen
options.headless = True

#geckodriver will be on the same path as selen.py
driver = webdriver.Firefox(executable_path='./geckodriver', options=options)


